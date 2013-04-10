from openpyxl import load_workbook
from time import time
import code

def getBest(database):
  highest = max(database.values())
  for key in database:
    if database[key] == highest:
      return key  

# ---------------------------------------------
# READ DATA FILE INTO MEMORY AS WORKBOOK
# ---------------------------------------------
files = [
  'gtd1993_1012dist.xlsx',
  'gtd_70to90_1012dist.xlsx',
  'gtd_91to11_1012dist.xlsx'
]

print 'Reading data file.'
t  = time()
ws = load_workbook(
  filename = '/Users/mukul/Desktop/gtd_201210dist/%s' % files[1], 
  use_iterators = True
).worksheets[0]
print 'Done. Time taken: %f secs.\n' % (time()-t)

# ---------------------------------------------
# BEGIN FILLING SCRIPT BUFFER
# ---------------------------------------------
print 'Obtaining lat/lng and country info.'

out_countries = open('countries.txt', 'w')
countries = {}

year = 1970
out_script = open('markers%s.js' % year, 'w')

latlngs = set()
geos = set()
t = time()

for row in ws.iter_rows():
  print [x.internal_value for x in row]
  raw_input()
  if row[0].internal_value == u'eventid': continue
  if row[1].internal_value > float(year):
    print 'For year %d, took %f seconds.' % (year, time()-t)
    print '%d total events; %d with lat/lon info, %d without' % (
      len(latlngs) + len(geos), len(latlngs), len(geos)
    )

    for (lat, lng) in latlngs:
      out_script.write('addMarkerAtLatLng(%s,%s);\n' % (lat, lng))
    for (city,country) in geos:
      out_script.write('addMarkerAtGeo("%s %s");\n' % (city, country))
    out_script.close()

    year += 1
    out_script = open('markers%s.js' % year, 'w')
    latlngs = set()
    geos = set()
    t = time()

  country = row[8].internal_value
  city = row[12].internal_value
  lat = row[13].internal_value
  lng = row[14].internal_value

  # ---------------------------------------------
  # OBTAINING AND RECORDING LAT/LNG COORDINATE INFO
  # ---------------------------------------------

  if row[1].internal_value == float(year):
    if type(lat) == float:
      latlngs.add(tuple([lat, lng]))
    elif city != None and '"' not in city and "'" not in city:
      geos.add(tuple([city, country]))
    else:
      continue

    # ---------------------------------------------
    # PROCESSING COUNTRY INFORMATION 
    # ---------------------------------------------
    try: countries[country] += 1
    except KeyError: countries[country] = 1


print 'For year %d, took %d seconds.' % (year, time()-t)
print '%d total events; %d with lat/lon info, %d without' % (
  len(latlngs) + len(geos), len(latlngs), len(geos)
)

for (lat, lng) in latlngs:
  out_script.write('addMarkerAtLatLng(%s,%s);\n' % (lat, lng))
for (city,country) in geos:
  out_script.write('addMarkerAtGeo("%s %s");\n' % (city, country))
out_script.close()
