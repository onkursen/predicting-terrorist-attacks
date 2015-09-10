from openpyxl import load_workbook
from time import time

# Read data file into memory as workbook
print 'Reading data file.'
t  = time()
ws = load_workbook(
  filename = raw_input("Enter path of GTD data file: ").strip(),
  use_iterators = True
).worksheets[0]
print 'Done. Time taken: %f secs.\n' % (time()-t)

inputs = []
outputs = []

# Get relevant values from database and put into input and output vectors for SVM classification
print 'Parsing database.'
t = time()

train = open('svm-train.txt', 'w')
test = open('svm-test.txt', 'w')

k = 0
for row in ws.iter_rows():
  if row[0].internal_value == u'eventid': continue

  iyear, imonth, iday, attack_type, targ_type, country = [row[j].internal_value for j in [1,2,3,28,34,7]]

  out_string = '%f %f %f %f %f %d\n' % (iyear, imonth, iday, attack_type, targ_type,country)

  if k % 2 == 0: train.write(out_string)
  else: test.write(out_string)
  k += 1

train.close()
test.close()

print 'Done. Time taken: %f secs.\n' % (time()-t)